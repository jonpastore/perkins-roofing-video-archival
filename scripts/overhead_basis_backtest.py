#!/usr/bin/env python3
"""Score the overhead-basis scenarios against what Tim ACTUALLY CHARGED on his own 30 homes.

The what-if in `overhead_basis_whatif.py` says how much each scenario MOVES a price. It cannot
say which one is RIGHT. This does: Tim's 2026-07-27 workbook carries, per home, his own day
counts AND his actual price per roof type, so each scenario can be scored against the number he
put in front of a customer.

Uses HIS day counts, never our day model — otherwise an error in the day estimate would be
scored as an error in the overhead basis, and Jon's 2026-07-31 mail already measured that the
days dominate ("an error in days is the same error in price, and it dwarfs the rate question").

⚠️ These are all Palm Beach / Martin / St Lucie homes, so FBC, and they are JUPITER's rates. This
scores the jupiter/naples basis question. It says nothing about Miami, which has its own burn and
no comparable price log.

Usage: DB_URL=… PYTHONPATH=. .venv/bin/python scripts/overhead_basis_backtest.py
"""
from __future__ import annotations

import json
import os
import statistics as st
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, text

from core.estimator import DailyOverheadSeries, QuoteInput, estimate
from core.pricing_config import load_config

SHEET = Path.home() / ("perkins-corpus/roofr-attachments/2026-07-27__Residential_OH_Calculator_"
                       "SLOPED_ONLY__with_stories_pitch_access_and_ACTUAL_PRICES.xlsx")

# column index (0-based) -> meaning, from the header row
C_EXISTING, C_STORIES, C_ACCESS = 4, 5, 7
C_SQ_SLOPED, C_SQ_FLAT = 8, 9
C_DEMO_D, C_SHINGLE_D, C_TILE_D, C_METAL_D, C_FLAT_D = 10, 11, 12, 13, 14
C_PRICE_SHINGLE, C_PRICE_TILE, C_PRICE_METAL = 15, 16, 17

# Which roof type each actual-price column is, and which day column feeds it.
KINDS = [
    ("shingle", C_PRICE_SHINGLE, C_SHINGLE_D, "dimensional_shingle"),
    ("tile",    C_PRICE_TILE,    C_TILE_D,    "13_tile"),
    ("metal",   C_PRICE_METAL,   C_METAL_D,   "standing_seam_metal"),
]

TIM_CREWS = 1.5      # his stated assumption, 2026-07-30
MEASURED_CREWS = 1.2  # scripts/recovery_identity_jupiter.py, charged job-days per working day
TIM_DAILY = 1470.0   # his stated Jupiter burn, 2026-07-30 (config carries 1400)


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def main() -> None:
    eng = create_engine(os.environ["DB_URL"])
    with eng.connect() as c:
        c.execute(text("set app.tenant_id='1'"))
        raw = c.execute(text("select config from pricing_configs "
                             "where branch='jupiter' and is_active")).scalar()
    live = raw if isinstance(raw, dict) else json.loads(raw)

    scenarios = {
        "1 today — series (his 4 rates)": dict(live),
        "2 branch basis, crews 1.5 (his)": {**live, "overhead_basis": "branch",
                                            "concurrent_crews": TIM_CREWS,
                                            "office_daily_overhead": TIM_DAILY},
        "3 branch basis, crews 1.2 (measured)": {**live, "overhead_basis": "branch",
                                                 "concurrent_crews": MEASURED_CREWS,
                                                 "office_daily_overhead": TIM_DAILY},
        "4 branch basis, crews 1.0 (today's default)": {**live, "overhead_basis": "branch",
                                                        "concurrent_crews": 1.0,
                                                        "office_daily_overhead": TIM_DAILY},
    }
    cfgs = {k: load_config(v) for k, v in scenarios.items()}

    ws = openpyxl.load_workbook(SHEET, data_only=True)[
        openpyxl.load_workbook(SHEET, data_only=True).sheetnames[0]]
    errs: dict[str, list[float]] = {k: [] for k in scenarios}
    n_cases = 0
    flat_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        sq = _num(row[C_SQ_SLOPED])
        if sq <= 0:
            continue
        flat_sq, flat_d = _num(row[C_SQ_FLAT]), _num(row[C_FLAT_D])
        if flat_sq > 0 and flat_d > 0:
            flat_rows += 1
        demo_d = _num(row[C_DEMO_D])
        existing = str(row[C_EXISTING] or "").strip().lower() or None
        access = str(row[C_ACCESS] or "").strip().upper().startswith("Y")

        for _label, price_col, day_col, roof_type in KINDS:
            actual, days = _num(row[price_col]), _num(row[day_col])
            if actual <= 0 or days <= 0:
                continue
            series = [DailyOverheadSeries(series=_series_for(roof_type), days=days)]
            # Demo AND flat both ride the demo_dry_in_flat series — Tim, 2026-07-24: "Demo &
            # Flat: $1,050 per day". Dropping the flat days (and the flat squares below) prices
            # a roof his actual price charged for, and reads as the engine under-quoting.
            if demo_d + flat_d > 0:
                series.append(DailyOverheadSeries(series="demo_dry_in_flat", days=demo_d + flat_d))
            q = QuoteInput(
                roof_type=roof_type, num_squares=sq, code_zone="FBC", slope_type="sloped",
                flat_squares=flat_sq,
                flat_roof_type="polyglass_sav_sap" if flat_sq > 0 else None,
                overhead_mode="daily", daily_series=series, access_difficult=access,
                existing_roof=existing if existing in
                {"none", "shingle", "tile", "metal", "flat"} else None,
            )
            n_cases += 1
            for name, cfg in cfgs.items():
                got = estimate(cfg, q)["project_total"]
                errs[name].append((got - actual) / actual)

    print(f"{n_cases} priced cases from Tim's own homes (his day counts, his actual prices)")
    print(f"{flat_rows} of his homes carry BOTH flat squares and flat days — the low-slope "
          f"day data the 'SLOPED ONLY' filename implies does not exist\n")
    print(f"    {'scenario':46} {'median err':>11} {'within 5%':>10} {'within 10%':>11} "
          f"{'|median|':>9}")
    for name, e in errs.items():
        w5 = sum(1 for x in e if abs(x) <= 0.05)
        w10 = sum(1 for x in e if abs(x) <= 0.10)
        print(f"    {name:46} {100*st.median(e):>10.1f}% {w5:>7}/{len(e)} {w10:>8}/{len(e)} "
              f"{100*st.median([abs(x) for x in e]):>8.1f}%")


def _series_for(roof_type: str) -> str:
    if "tile" in roof_type:
        return "tile"
    if "shingle" in roof_type:
        return "shingle"
    return "metal"


if __name__ == "__main__":
    main()
